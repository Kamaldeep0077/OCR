#image_path = "/Users/kamaldeep/Desktop/Work/LangchainModels/test/Screenshot 2025-05-14 at 12.30.14 AM.png"
import cv2
import numpy as np
import easyocr
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ─── Helpers ─────────────────────────────────────────────────────────────

def cluster_positions(vals, tol):
    """Sort & merge sorted values into cluster centers within tol."""
    centers = []
    for v in sorted(vals):
        if not centers or abs(v - centers[-1]) > tol:
            centers.append(float(v))
        else:
            centers[-1] = (centers[-1] + v) / 2.0
    return centers

def avg_color_argb(bgr_roi):
    """Compute average BGR color and return ARGB hex for Excel (FFRRGGBB)."""
    b, g, r = cv2.mean(bgr_roi)[:3]
    return f"FF{int(r):02X}{int(g):02X}{int(b):02X}"


# ─── 1) LOAD IMAGE ──────────────────────────────────────────────────────

# handle weird spaces/unicode in filename
cands = glob.glob("/Users/kamaldeep/Desktop/Work/LangchainModels/test/Screenshot*AM.jpg")
if not cands:
    raise FileNotFoundError("No screenshot matched under test/")
img_path = cands[0]
img = cv2.imread(img_path)
if img is None:
    raise ValueError(f"Could not load image at {img_path}")

# keep a color copy and make a grayscale for line detection
gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
inv  = cv2.bitwise_not(gray)
bw   = cv2.adaptiveThreshold(inv, 255,
                             cv2.ADAPTIVE_THRESH_MEAN_C,
                             cv2.THRESH_BINARY, 15, -2)

# ─── 2) DETECT TABLE LINES ───────────────────────────────────────────────

# horizontal lines
horiz = bw.copy()
h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT,
                                     (bw.shape[1]//15, 1))
horiz = cv2.erode (horiz, h_kernel, iterations=1)
horiz = cv2.dilate(horiz, h_kernel, iterations=1)

# vertical lines
vert = bw.copy()
v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT,
                                     (1, bw.shape[0]//15))
vert = cv2.erode (vert, v_kernel, iterations=1)
vert = cv2.dilate(vert, v_kernel, iterations=1)

# find all intersection points of grid
inter = cv2.bitwise_and(horiz, vert)
pts   = cv2.findNonZero(inter)
if pts is None:
    raise RuntimeError("No table intersections detected!")
pts = pts.reshape(-1, 2)
all_x = [p[0] for p in pts]
all_y = [p[1] for p in pts]

# cluster into unique grid lines
x_coords = cluster_positions(all_x, tol=10)
y_coords = cluster_positions(all_y, tol=10)

# number of cells
n_cols = len(x_coords) - 1
n_rows = len(y_coords) - 1

# ─── 3) OCR SETUP ────────────────────────────────────────────────────────

reader = easyocr.Reader(['en'])

# ─── 4) BUILD TABLE, DETECT MERGES ──────────────────────────────────────

table  = [[''] * n_cols for _ in range(n_rows)]
colors = [['FFFFFFFF'] * n_cols for _ in range(n_rows)]
merges = []  # list of (r1,c1,r2,c2)

for r in range(n_rows):
    for c in range(n_cols):
        # grid cell bounds
        x1, x2 = int(x_coords[c]),   int(x_coords[c+1])
        y1, y2 = int(y_coords[r]),   int(y_coords[r+1])

        # detect vertical span by checking for missing horizontal border
        span_r = 1
        while (r + span_r) < n_rows:
            yb = int(y_coords[r + span_r])
            xm = (x1 + x2)//2
            # if there's a horizontal line at yb,xm, stop
            if horiz[yb, xm] > 0:
                break
            span_r += 1

        # detect horizontal span by checking for missing vertical border
        span_c = 1
        while (c + span_c) < n_cols:
            xb = int(x_coords[c + span_c])
            ym = (y1 + y2)//2
            if vert[ym, xb] > 0:
                break
            span_c += 1

        # record merge region
        merges.append((r, c, r + span_r - 1, c + span_c - 1))

        # OCR the full merged region
        X2 = int(x_coords[c + span_c])
        Y2 = int(y_coords[r + span_r])
        roi_gray = gray[y1:Y2, x1:X2]
        txts = reader.readtext(roi_gray, detail=0)
        table[r][c] = " ".join(txts).strip()

        # average color from original color image
        colors[r][c] = avg_color_argb(img[y1:Y2, x1:X2])


# ─── 5) WRITE TO EXCEL ──────────────────────────────────────────────────

wb = Workbook()
ws = wb.active

# A) Write every cell (so no MergedCell errors)
for r in range(n_rows):
    for c in range(n_cols):
        val = table[r][c] or ""
        fc  = colors[r][c]
        cell = ws.cell(row=r+1, column=c+1, value=val)
        cell.fill = PatternFill(start_color=fc,
                                end_color=fc,
                                fill_type="solid")

# B) Apply merges, skip 1×1 duplicates
seen = set()
for r1, c1, r2, c2 in merges:
    if (r1,c1,r2,c2) in seen or (r1==r2 and c1==c2):
        continue
    seen.add((r1,c1,r2,c2))
    ws.merge_cells(start_row=r1+1, start_column=c1+1,
                   end_row=r2+1,   end_column=c2+1)

# C) Save
out = "/Users/kamaldeep/Desktop/Work/LangchainModels/test/ocr_colored_output1.xlsx"
wb.save(out)
print("Done →", out)
